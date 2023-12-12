import openpyxl
import random
from ..models import Beer,User
from .. import db

def init_database():
    dataframe = openpyxl.load_workbook("Beerco_kiosk_data.xlsx")
    dataframe1 = dataframe.active
    d = 0
    p = 5
    sp = 1
    sf = 1234
    for row in dataframe1.iter_rows(2, dataframe1.max_row):
        new_beer = Beer(id=d,price=p,name=str(row[2].value),brewer=str(row[3].value),region=str(row[4].value),
                        country=str(row[5].value),state=str(row[6].value),tempBeerStyle1=str(row[8].value),
                        tempBeerStyle2=str(row[9].value),primaryBeerStyle=str(row[10].value),primaryBeerStyleDesc=str(row[11].value),
                        specificBeerStyle=str(row[12].value),specificBeerStyleDesc=str(row[13].value),foodPairing=str(row[14].value),oz=str(row[15].value),
                        ml=str(row[16].value),container=str(row[17].value),abv=str(row[19].value),ibu=str(row[20].value),
                        srm=str(row[21].value),malt=str(row[22].value),hops=str(row[23].value),cals=str(row[25].value),
                        desc=str(row[26].value),special=sp,searchFreq=sf)
        db.session.add(new_beer)
        if d >= 3:
            sp = 0
            p = float(random.randint(10, 15))
        sf = random.randint(0, 1000)
        d += 1

    admin = User(role = 1, name="Admin", password=hash("password", "salt"), salt="salt")
    retailer = User(role = 2, name="Retailer", password=hash("password","hello"), salt="hello")
    kiosk = User(role = 3, name="Kiosk", password=hash("password","tang"), salt="tang")
    
    db.session.add(admin)
    db.session.add(retailer)
    db.session.add(kiosk)
    db.session.commit()
    
    print('DB initialized!')

def hash(string, salt):
    pw = bytearray(string+salt, encoding='utf-8')
    copy = bytearray()
    for c in pw:
        x = chr((c >> 1) + 65).encode('utf-8')
        copy.extend(x)
    return copy.decode('utf-8')    

def comparePasswords(password, user):
    hashedPassword =  hash(password, user.salt)
    if hashedPassword == user.password:
        return True
    return False

def findTotal(cart):
    total = 0.0
    for item in cart:
        total += float(item.price)
    return total